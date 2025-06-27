from torch.cuda.amp import GradScaler
import sys
import os

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(project_root)
from model.model_ensamble import au_graph_mer
from utils.datasset import MEDataset_161_landmark_7class_emo, AugmentedDataset, \
    MEDataset_161_landmark_7class_emo_triple_frames
from utils.calculate_metric import recognition_evaluation_7class, au_evaluate_12
from sklearn.model_selection import KFold
import argparse
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
from openpyxl import Workbook
from torch.utils.data import DataLoader


def initialize_weights(model):
    for m in model.modules():
        if isinstance(m, nn.Conv2d) or isinstance(m, nn.Linear):
            nn.init.kaiming_normal_(m.weight)
            if m.bias is not None:
                nn.init.constant_(m.bias, 0)


def main(args):
    df = pd.read_excel(args.excel_path, engine='openpyxl', dtype={'Subject': str, 'Filename': str})
    unique_subjects = df['Subject'].dropna().unique()
    label_mapping = {1: 0, 2: 1, 4: 2, 5: 3, 6: 4, 7: 5, 9: 6, 10: 7, 12: 8, 14: 9, 15: 10, 17: 11}
    device = torch.device(f'cuda:{args.cuda_device}' if torch.cuda.is_available() else 'cpu')
    sigmoid = nn.Sigmoid()
    au_criterion = nn.BCEWithLogitsLoss().to(device)
    emo_criterion = nn.CrossEntropyLoss().to(device)
    kf = KFold(n_splits=10, shuffle=True, random_state=1)
    wb = Workbook()
    wb.remove(wb.active)
    for fold, (train_index, test_index) in enumerate(kf.split(unique_subjects)):
        print(f"Fold {fold + 1}")
        test_subjects = unique_subjects[test_index]
        test_subjects_list = list(test_subjects)
        print(f"Test subjects: {test_subjects_list}")
        train_set = MEDataset_161_landmark_7class_emo(df, label_mapping, args.data_path, args.dataset, device, True,
                                                      test_subjects_list)
        test_set = MEDataset_161_landmark_7class_emo(df, label_mapping, args.data_path, args.dataset, device, False,
                                                     test_subjects_list)
        # train_set = MEDataset_161_landmark_7class_emo_triple_frames(df, label_mapping, args.data_path, args.dataset, device, True,
        #                                               test_subjects_list)
        # test_set = MEDataset_161_landmark_7class_emo_triple_frames(df, label_mapping, args.data_path, args.dataset, device, False,
        #                                              test_subjects_list)
        augmented_dataset = AugmentedDataset(train_set)
        train_loader = DataLoader(augmented_dataset, batch_size=args.batch_size, shuffle=True)
        test_loader = DataLoader(test_set, batch_size=args.batch_size)
        model = au_graph_mer('au_graph_mer_softmax', device, out_channels=7).to(device)
        model.apply(initialize_weights)
        scaler = GradScaler()
        optimizer = optim.AdamW(filter(lambda p: p.requires_grad, model.parameters()), lr=args.learning_rate)
        for epoch in range(args.epochs):
            percentage = (epoch + 1) / args.epochs
            args.ratio = args.ratio * (1 - percentage)
            model.train()
            running_loss = 0.0
            all_au_predicts, all_emo_predicts, all_au_labels, all_emo_labels = [], [], [], []
            for i, data in enumerate(train_loader, 0):
                global_features, au_features, au_labels, emo_labels = data
                global_features, au_features, au_labels, emo_labels = global_features.to(
                    device).float(), au_features.to(device).float(), au_labels.to(device).float(), emo_labels.to(
                    device)
                optimizer.zero_grad()
                with torch.cuda.amp.autocast():
                    emo_predicts, au_predicts, auxiliary_au_predicts = model(global_features, au_features, args.ratio)
                    au_loss = au_criterion(au_predicts, au_labels) + au_criterion(auxiliary_au_predicts, au_labels)
                    emo_loss = emo_criterion(emo_predicts, emo_labels.long())
                    total_loss = (1 - args.gamma) * emo_loss + args.gamma * au_loss
                all_emo_predicts.extend(torch.max(emo_predicts, 1)[1].tolist())
                all_emo_labels.extend(emo_labels.long().tolist())
                all_au_predicts.append(au_predicts)
                all_au_labels.append(au_labels)
                scaler.scale(total_loss).backward()
                scaler.step(optimizer)
                scaler.update()
                running_loss += total_loss.item()
                # scheduler.step()
            if (epoch + 1) % 5 == 0:
                train_acc = sum(x == y for x, y in zip(all_emo_predicts, all_emo_labels)) / len(all_emo_labels)

                print(f'lr:{args.learning_rate}, epoch:{epoch + 1}, train_acc：{train_acc}')
                model.eval()
                with torch.no_grad():
                    test_loss = 0.0
                    for data in test_loader:
                        global_features, au_features, au_labels, emo_labels = data
                        global_features, au_features, au_labels, emo_labels = global_features.to(
                            device).float(), au_features.to(device).float(), au_labels.to(
                            device).float(), emo_labels.to(
                            device)
                        emo_predicts, au_predicts, auxiliary_au_predicts = model(global_features, au_features,
                                                                                 args.ratio)
                        au_loss = au_criterion(au_predicts, au_labels) + au_criterion(auxiliary_au_predicts, au_labels)
                        emo_loss = emo_criterion(emo_predicts, emo_labels.long())
                        test_loss = (1 - args.gamma) * emo_loss + args.gamma * au_loss
                    print(f'lr:{args.learning_rate}, epoch:{epoch + 1}, train_loss:{running_loss}, test_loss:{test_loss}')
            if (epoch + 1) % 1 == 0:
                model.eval()
                with torch.no_grad():
                    all_au_predicts, all_emo_predicts, all_au_labels, all_emo_labels = [], [], [], []
                    for data in test_loader:
                        global_features, au_features, au_labels, emo_labels = data
                        global_features, au_features, au_labels, emo_labels = global_features.to(
                            device).float(), au_features.to(device).float(), au_labels.to(
                            device).float(), emo_labels.to(
                            device)
                        emo_predicts, au_predicts, auxiliary_au_predicts = model(global_features, au_features,
                                                                                 args.ratio)
                        all_emo_predicts.extend(torch.max(emo_predicts, 1)[1].tolist())
                        all_emo_labels.extend(emo_labels.long().tolist())
                        all_au_predicts.append(au_predicts)
                        all_au_labels.append(au_labels)
                    all_au_predicts = torch.cat(all_au_predicts, dim=0)
                    all_au_labels = torch.cat(all_au_labels, dim=0)
                    f1_list, recall_list, emo_UF1, emo_UAR = recognition_evaluation_7class(all_emo_labels,
                                                                                           all_emo_predicts, show=False)
                    cur_acc = sum(x == y for x, y in zip(all_emo_predicts, all_emo_labels)) / len(all_emo_predicts)
                    print(f'lr:{args.learning_rate}, epoch:{epoch + 1}, test_acc：{cur_acc}')
                    test_au_recall_list, test_au_f1_score_list, test_au_accuracy_list = au_evaluate_12(
                        sigmoid(all_au_predicts), all_au_labels)
                    final_list = [fold, f1_list[0], f1_list[1], f1_list[2], f1_list[3], f1_list[4], f1_list[5],
                                  f1_list[6], emo_UF1, recall_list[0], recall_list[1], recall_list[2],
                                  recall_list[3], recall_list[4], recall_list[5], recall_list[6], emo_UAR,
                                  cur_acc, test_au_accuracy_list[0], test_au_accuracy_list[1],
                                  test_au_accuracy_list[2],
                                  test_au_accuracy_list[3], test_au_accuracy_list[4], test_au_accuracy_list[5],
                                  test_au_accuracy_list[6], test_au_accuracy_list[7],
                                  test_au_accuracy_list[8], test_au_accuracy_list[9], test_au_accuracy_list[10],
                                  test_au_accuracy_list[11], test_au_f1_score_list[0], test_au_f1_score_list[1],
                                  test_au_f1_score_list[2],
                                  test_au_f1_score_list[3], test_au_f1_score_list[4], test_au_f1_score_list[5],
                                  test_au_f1_score_list[6], test_au_f1_score_list[7],
                                  test_au_f1_score_list[8], test_au_f1_score_list[9], test_au_f1_score_list[10],
                                  test_au_f1_score_list[11], test_au_recall_list[0], test_au_recall_list[1],
                                  test_au_recall_list[2],
                                  test_au_recall_list[3], test_au_recall_list[4], test_au_recall_list[5],
                                  test_au_recall_list[6], test_au_recall_list[7],
                                  test_au_recall_list[8], test_au_recall_list[9], test_au_recall_list[10],
                                  test_au_recall_list[11], ','.join(map(str, all_emo_predicts)),
                                  ','.join(map(str, all_emo_labels))]
                    sheet_name = f"Epoch_{epoch + 1}"
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=sheet_name)
                        ws.append(
                            ["Fold", "anger_F1", "contempt_F1", "disgust_F1", "fear_F1", "happiness_F1", "sadness_F1",
                             "surprise_F1", "U_Emo_Average_F1", "anger_Recall", "contempt_Recall", "disgust_Recall",
                             "fear_Recall", "happiness_Recall", "sadness_Recall", "surprise_Recall",
                             "U_Emo_Average_Recall",
                             "Emo_Acc", "AU1_Acc", "AU2_Acc", "AU4_Acc",
                             "AU5_Acc", "AU6_Acc", "AU7_Acc", "AU9_Acc", "AU10_Acc", "AU12_Acc", "AU14_Acc",
                             "AU15_Acc", "AU17_Acc",
                             "AU1_F1", "AU2_F1", "AU4_F1", "AU5_F1", "AU6_F1", "AU7_F1",
                             "AU9_F1",
                             "AU10_F1", "AU12_F1", "AU14_F1", "AU15_F1", "AU17_F1",
                             "AU1_Recall", "AU2_Recall",
                             "AU4_Recall", "AU5_Recall", "AU6_Recall", "AU7_Recall", "AU9_Recall", "AU10_Recall",
                             "AU12_Recall",
                             "AU14_Recall", "AU15_Recall", "AU17_Recall", "emo_predicts", "emo_labels"
                             ])
                    else:
                        ws = wb[sheet_name]
                    ws.append(final_list)
        print('Finished Training')

    wb.save(args.output_path)
    print('Excel Saved！')


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Train and evaluate AU MTL Network")
    parser.add_argument('--excel_path', type=str, default='../data_coding/DFME_data_coding.xlsx',
                        help='Path to the Excel file')
    parser.add_argument('--data_path', type=str, default='../data/dfme_augment', help='Path to the data directory')
    parser.add_argument('--dataset', type=str, default='DFME', choices=['CASME2', 'CASME^3', 'SAMM', 'DFME'],
                        help='Choose dataset to train and validate')
    parser.add_argument('--output_path', type=str,
                        default='',
                        help='Path to save the output Excel file')
    parser.add_argument('--epochs', type=int, default=30, help='Number of epochs to train')
    parser.add_argument('--ratio', type=float, default=1.0, help='The initial proportion of prior knowledge')
    parser.add_argument('--batch_size', type=int, default=256, help='Batch size for training and testing')
    parser.add_argument('--learning_rate', type=float, default=0.00005, help='Learning rate for the optimizer')
    parser.add_argument('--gamma', type=float, default=0.8, help='au loss ratio')
    parser.add_argument('--cuda_device', type=int, default=1, help='CUDA device number')
    parser.add_argument('--pre_trained', type=bool, default=False, help='Choose whether pretrained au model')
    args = parser.parse_args()
    main(args)
